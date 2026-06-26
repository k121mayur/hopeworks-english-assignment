"""Reports router — attendance, progress, difficult words, dashboard."""

from datetime import date, timedelta
from typing import Optional
from io import BytesIO

from fastapi import APIRouter, Depends, Query, HTTPException
from fastapi.responses import StreamingResponse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from ..auth import require_admin
from ..database import get_db
from ..models import Class, ClassStudent, Story, Submission, User, WordError
from ..schemas import (
    AttendanceRecord,
    DashboardMetrics,
    DifficultWord,
    ProgressRecord,
)

router = APIRouter(
    prefix="/api/reports",
    tags=["Reports"],
    dependencies=[Depends(require_admin)],
)


@router.get("/attendance", response_model=list[AttendanceRecord])
async def attendance_report(
    class_id: Optional[int] = Query(None),
    start_date: Optional[date] = Query(None),
    end_date: Optional[date] = Query(None),
    db: AsyncSession = Depends(get_db),
):
    """Daily attendance — which students submitted readings."""
    # Build student query
    student_q = select(User).where(User.role == "student")
    if class_id:
        student_q = student_q.join(ClassStudent).where(ClassStudent.class_id == class_id)
    student_result = await db.execute(student_q)
    students = student_result.scalars().all()

    # Build story query
    story_q = select(Story)
    if start_date:
        story_q = story_q.where(Story.created_date >= start_date)
    if end_date:
        story_q = story_q.where(Story.created_date <= end_date)
    story_q = story_q.order_by(Story.created_date.desc())
    story_result = await db.execute(story_q)
    stories = story_result.scalars().all()

    # Find all submissions as a set for fast lookup
    sub_result = await db.execute(
        select(Submission.student_id, Submission.story_id)
    )
    submitted_set = {(row[0], row[1]) for row in sub_result.all()}

    records = []
    for story in stories:
        for student in students:
            records.append(
                AttendanceRecord(
                    student_name=student.name,
                    student_email=student.email,
                    date=story.created_date,
                    story_id=story.id,
                    story_title=story.title,
                    submitted=(student.id, story.id) in submitted_set,
                )
            )
    return records


@router.get("/progress/{student_id}", response_model=list[ProgressRecord])
async def student_progress(student_id: str, db: AsyncSession = Depends(get_db)):
    """Accuracy scores over time for one student (using max score per story)."""
    result = await db.execute(
        select(
            Story.title,
            func.max(Submission.accuracy_score).label("max_score"),
            func.min(Submission.created_at).label("first_submitted")
        )
        .join(Submission, Submission.story_id == Story.id)
        .where(Submission.student_id == student_id)
        .group_by(Story.id, Story.title)
        .order_by(func.min(Submission.created_at))
    )
    rows = result.all()
    return [
        ProgressRecord(
            date=row.first_submitted.date(),
            story_title=row.title,
            accuracy_score=row.max_score or 0.0,
        )
        for row in rows
    ]


@router.get("/difficult-words", response_model=list[DifficultWord])
async def difficult_words(
    limit: int = Query(50, le=200),
    db: AsyncSession = Depends(get_db),
):
    """Most-mispronounced words across the platform."""
    result = await db.execute(
        select(
            WordError.expected_word,
            func.count().label("error_count"),
        )
        .where(WordError.expected_word != "")
        .group_by(WordError.expected_word)
        .order_by(func.count().desc())
        .limit(limit)
    )
    return [
        DifficultWord(word=row[0], error_count=row[1]) for row in result.all()
    ]


@router.get("/dashboard", response_model=DashboardMetrics)
async def dashboard_metrics(db: AsyncSession = Depends(get_db)):
    """Aggregate metrics for the admin dashboard."""
    # Subquery: max accuracy_score per student per story
    subquery = (
        select(
            Submission.student_id,
            Submission.story_id,
            func.max(Submission.accuracy_score).label("max_score")
        )
        .group_by(Submission.student_id, Submission.story_id)
        .subquery()
    )
    avg_acc_res = await db.execute(select(func.avg(subquery.c.max_score)))
    avg_acc = avg_acc_res.scalar()

    total_subs = await db.execute(select(func.count()).select_from(Submission))
    total_students = await db.execute(
        select(func.count()).select_from(User).where(User.role == "student")
    )
    total_stories = await db.execute(select(func.count()).select_from(Story))

    return DashboardMetrics(
        average_accuracy=round(avg_acc or 0, 2),
        total_submissions=total_subs.scalar() or 0,
        total_students=total_students.scalar() or 0,
        total_stories=total_stories.scalar() or 0,
    )


@router.get("/class-report")
async def class_report(
    class_id: int = Query(...),
    range_type: str = Query("last_week"),  # "last_week" or "last_month"
    download: bool = Query(False),
    db: AsyncSession = Depends(get_db),
):
    today = date.today()
    if range_type == "last_week":
        start_date = today - timedelta(days=7)
    elif range_type == "last_month":
        start_date = today - timedelta(days=30)
    else:
        start_date = today - timedelta(days=7)
    end_date = today

    # Fetch class
    cls_res = await db.execute(select(Class).where(Class.id == class_id))
    cls = cls_res.scalar_one_or_none()
    if not cls:
        raise HTTPException(status_code=404, detail="Class not found")

    # Fetch students
    student_res = await db.execute(
        select(User)
        .join(ClassStudent)
        .where(ClassStudent.class_id == class_id, User.role == "student")
        .order_by(User.name)
    )
    students = student_res.scalars().all()

    # Fetch stories
    story_res = await db.execute(
        select(Story)
        .where(Story.created_date >= start_date, Story.created_date <= end_date)
        .order_by(Story.created_date.desc())
    )
    stories = story_res.scalars().all()

    student_ids = [s.id for s in students]
    story_ids = [st.id for st in stories]

    # Fetch submissions
    submissions = []
    if student_ids and story_ids:
        sub_res = await db.execute(
            select(Submission)
            .where(Submission.student_id.in_(student_ids), Submission.story_id.in_(story_ids))
        )
        submissions = sub_res.scalars().all()

    # Group submissions
    # Map (student_id, story_id) -> list of accuracy_scores
    sub_map = {}
    for sub in submissions:
        key = (sub.student_id, sub.story_id)
        if key not in sub_map:
            sub_map[key] = []
        sub_map[key].append(sub.accuracy_score)

    records = []
    total_present = 0
    total_possible = len(students) * len(stories)

    for story in stories:
        for student in students:
            key = (student.id, story.id)
            scores = sub_map.get(key, [])
            attempts = len(scores)
            max_score = max(scores) if scores else 0.0

            # Present if submissions exist and best score >= 80
            is_present = len(scores) > 0 and max_score >= 80.0
            status = "Present" if is_present else "Absent"
            if is_present:
                total_present += 1

            records.append({
                "student_name": student.name,
                "student_email": student.email,
                "date": story.created_date.strftime("%Y-%m-%d"),
                "story_title": story.title,
                "attempts": attempts,
                "max_score": round(max_score, 2),
                "status": status
            })

    attendance_rate = (total_present / total_possible * 100) if total_possible > 0 else 0.0

    if download:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Class Attendance Report"

        # Colors Matching Slate/Premium Theme
        navy_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
        light_gray_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
        green_fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")  # Present
        red_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")  # Absent

        white_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        bold_font = Font(name="Arial", size=11, bold=True)
        regular_font = Font(name="Arial", size=10)
        title_font = Font(name="Arial", size=16, bold=True, color="0F172A")

        thin_border = Border(
            left=Side(style='thin', color='E2E8F0'),
            right=Side(style='thin', color='E2E8F0'),
            top=Side(style='thin', color='E2E8F0'),
            bottom=Side(style='thin', color='E2E8F0')
        )

        # Title
        ws.merge_cells("A1:G1")
        ws["A1"] = f"Class Attendance Report: {cls.class_name}"
        ws["A1"].font = title_font
        ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 40

        # Subtitle Info
        ws["A2"] = "Date Range:"
        ws["A2"].font = bold_font
        ws["B2"] = f"{start_date} to {end_date} ({range_type.replace('_', ' ').title()})"
        ws["B2"].font = regular_font

        ws["A3"] = "Attendance Rate:"
        ws["A3"].font = bold_font
        ws["B3"] = f"{round(attendance_rate, 1)}%"
        ws["B3"].font = regular_font

        ws["D2"] = "Total Students:"
        ws["D2"].font = bold_font
        ws["E2"] = len(students)
        ws["E2"].font = regular_font

        ws["D3"] = "Total Stories:"
        ws["D3"].font = bold_font
        ws["E3"] = len(stories)
        ws["E3"].font = regular_font

        ws.row_dimensions[2].height = 20
        ws.row_dimensions[3].height = 20

        # Empty row
        ws.row_dimensions[4].height = 15

        # Table Headers
        headers = [
            "Student Name",
            "Email Address",
            "Story Assignment Date",
            "Story Title",
            "Attempts",
            "Best Score (%)",
            "Attendance Status",
        ]
        ws.row_dimensions[5].height = 26

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col_idx, value=header)
            cell.font = white_font
            cell.fill = navy_fill
            cell.alignment = Alignment(
                horizontal="center" if col_idx > 2 else "left", vertical="center"
            )
            cell.border = thin_border

        # Write Data
        for row_idx, rec in enumerate(records, 6):
            ws.row_dimensions[row_idx].height = 22

            c1 = ws.cell(row=row_idx, column=1, value=rec["student_name"])
            c2 = ws.cell(row=row_idx, column=2, value=rec["student_email"])
            c3 = ws.cell(row=row_idx, column=3, value=rec["date"])
            c4 = ws.cell(row=row_idx, column=4, value=rec["story_title"])
            c5 = ws.cell(row=row_idx, column=5, value=rec["attempts"])
            c6 = ws.cell(row=row_idx, column=6, value=rec["max_score"])
            c7 = ws.cell(row=row_idx, column=7, value=rec["status"])

            # Fonts & borders
            for cell in [c1, c2, c3, c4, c5, c6, c7]:
                cell.font = regular_font
                cell.border = thin_border

            # Alignments
            c1.alignment = Alignment(horizontal="left", vertical="center")
            c2.alignment = Alignment(horizontal="left", vertical="center")
            c3.alignment = Alignment(horizontal="center", vertical="center")
            c4.alignment = Alignment(horizontal="left", vertical="center")
            c5.alignment = Alignment(horizontal="center", vertical="center")
            c6.alignment = Alignment(horizontal="right", vertical="center")
            c7.alignment = Alignment(horizontal="center", vertical="center")

            # Formatting numbers
            c6.number_format = '0.00'

            # Highlight status
            if rec["status"] == "Present":
                c7.fill = green_fill
                c7.font = Font(name="Arial", size=10, bold=True, color="065F46")  # dark green
            else:
                c7.fill = red_fill
                c7.font = Font(name="Arial", size=10, bold=True, color="991B1B")  # dark red

            # Striping
            if row_idx % 2 == 1:
                for cell in [c1, c2, c3, c4, c5, c6]:
                    cell.fill = light_gray_fill

        # Auto-adjust column widths
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.row > 4 and cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

        # Save to buffer
        out = BytesIO()
        wb.save(out)
        out.seek(0)

        filename = f"class_{class_id}_report_{range_type}_{today.strftime('%Y%m%d')}.xlsx"
        return StreamingResponse(
            out,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )

    return {
        "class_id": class_id,
        "class_name": cls.class_name,
        "range_type": range_type,
        "start_date": start_date,
        "end_date": end_date,
        "attendance_rate": round(attendance_rate, 2),
        "total_possible": total_possible,
        "total_present": total_present,
        "total_students": len(students),
        "total_stories": len(stories),
        "records": records,
    }
